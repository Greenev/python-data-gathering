from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font
import re

cols = ('Название', 'ФИО Руководителя', 'Контакты', 'Месяц', 'Адрес', 'ИНН')
letters = ('A', 'B', 'C', 'D', 'E', 'F')
widths = (35, 32, 45, 10, 34, 11)

def excel_parsing(first, last, path, objects):
    #columns selecting
    table_name = 'B'
    table_month = 'M'
    table_inn = 'G'
    table_adress = 'E'
    #.xlsx source file loading
    wb = load_workbook(path)
    ws = wb.active
    #duplicats checking by name
    names = []
    for number in range(first, last):
        name = re.sub('(.*ТСЖ)|(.*Товарищество со+бственников жилья)', 'ТСЖ', ws[table_name + str(number)].value)
        name = re.sub('(.*ЖСК)|(.*Жилищно-строительный ко+п+ератив)', 'ЖСК', name)
        if re.findall('(ТСЖ)|(ЖСК)', name) and not name in names:
            names.append(name)
            objects.append({cols[0] : name,
                            cols[5] : ws[table_inn + str(number)].value,
                            cols[3] : ws[table_month + str(number)].value,
                            cols[4] : re.sub('.*Санкт-Петербург,', '', ws[table_adress + str(number)].value).lstrip()})

def db_create(db_path, objects):
    #duplicats checking by contact info
    contacts = []
    #.xlsx target file creating
    wb_db = Workbook()
    ws_db = wb_db.active
    ws_db.append(cols)
    #bold titles
    for col in range(1, 7):
        ws_db.cell(row = 1, column = col).font = Font(bold = True)
    #columns width setting
    width_set = dict(zip(letters, widths))    
    for i in width_set:
        ws_db.column_dimensions[i].width = width_set[i]
    #columns filling
    for row in objects:
        norml = re.sub('-| ', '', row[cols[2]])
        if norml not in contacts:
            contacts.append(norml)
            ws_db.append([row[i] for i in cols])
        
    ws_db.column_dimensions['F'].hidden = True
    wb_db.save(filename = db_path)


if __name__ == '__main__':
    print('This script provides the .xlsx files processing functions')
